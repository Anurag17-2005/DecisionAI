#!/usr/bin/env python3
"""
Document Processor for Vector Database
Extracts text from PDF, DOCX, XLSX files and prepares them for indexing
"""

import os
from typing import List, Dict
from pathlib import Path

# PDF processing
from pypdf import PdfReader

# DOCX processing
from docx import Document

# XLSX processing
import openpyxl

class DocumentProcessor:
    """Extract text from various document formats"""
    
    def __init__(self, docs_directory: str = "business_documents"):
        self.docs_directory = docs_directory
    
    def extract_text_from_pdf(self, filepath: str) -> str:
        """Extract text from PDF file"""
        try:
            reader = PdfReader(filepath)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            print(f"Error reading PDF {filepath}: {e}")
            return ""
    
    def extract_text_from_docx(self, filepath: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(filepath)
            text = ""
            for para in doc.paragraphs:
                text += para.text + "\n"
            return text.strip()
        except Exception as e:
            print(f"Error reading DOCX {filepath}: {e}")
            return ""
    
    def extract_text_from_xlsx(self, filepath: str) -> str:
        """Extract text from XLSX file"""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            text = ""
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n=== Sheet: {sheet_name} ===\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            
            return text.strip()
        except Exception as e:
            print(f"Error reading XLSX {filepath}: {e}")
            return ""
    
    def process_document(self, filepath: str) -> Dict[str, str]:
        """Process a single document and return metadata + text"""
        file_ext = Path(filepath).suffix.lower()
        filename = Path(filepath).name
        
        if file_ext == '.pdf':
            text = self.extract_text_from_pdf(filepath)
        elif file_ext == '.docx':
            text = self.extract_text_from_docx(filepath)
        elif file_ext == '.xlsx':
            text = self.extract_text_from_xlsx(filepath)
        else:
            text = ""
        
        return {
            'filename': filename,
            'filepath': filepath,
            'filetype': file_ext,
            'text': text
        }
    
    def process_all_documents(self) -> List[Dict[str, str]]:
        """Process all documents in the directory"""
        documents = []
        
        if not os.path.exists(self.docs_directory):
            print(f"Directory {self.docs_directory} not found!")
            return documents
        
        for filename in os.listdir(self.docs_directory):
            filepath = os.path.join(self.docs_directory, filename)
            
            if os.path.isfile(filepath) and filename.lower().endswith(('.pdf', '.docx', '.xlsx')):
                print(f"Processing: {filename}")
                doc = self.process_document(filepath)
                if doc['text']:
                    documents.append(doc)
                    print(f"  ✅ Extracted {len(doc['text'])} characters")
                else:
                    print(f"  ⚠️ No text extracted")
        
        return documents
    
    def chunk_text(self, text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """Split text into overlapping chunks for better retrieval"""
        chunks = []
        start = 0
        text_len = len(text)
        
        while start < text_len:
            end = start + chunk_size
            chunk = text[start:end]
            
            # Try to break at sentence boundary
            if end < text_len:
                last_period = chunk.rfind('.')
                last_newline = chunk.rfind('\n')
                break_point = max(last_period, last_newline)
                
                if break_point > chunk_size * 0.5:  # Only break if not too early
                    chunk = chunk[:break_point + 1]
                    end = start + break_point + 1
            
            chunks.append(chunk.strip())
            start = end - overlap
        
        return chunks

if __name__ == "__main__":
    processor = DocumentProcessor()
    docs = processor.process_all_documents()
    print(f"\n✅ Processed {len(docs)} documents")
